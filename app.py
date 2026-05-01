import io
import re
import sys
import time
import unicodedata
from dataclasses import dataclass
from typing import Optional

import pandas as pd
import requests

try:
    import streamlit as st
except ImportError:
    st = None

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
    from playwright.sync_api import sync_playwright
    PLAYWRIGHT_AVAILABLE = True
except ImportError:
    PlaywrightTimeoutError = Exception
    sync_playwright = None
    PLAYWRIGHT_AVAILABLE = False


# --------------------------------------------------------------------------- #
# Constants
# --------------------------------------------------------------------------- #
FALSE_POSITIVE_THRESHOLD = 25
REBNY_DIRECTORY_URL = "https://www.rebny.com/members/"
REPUBLICAN_PARTY_CODES = {"REP"}
REPUBLICAN_COMMITTEE_KEYWORDS = [
    "republican", "rnc", "nrcc", "nrsc", "gop", "maga",
    "america first", "trump", "conservative", "right to rise",
    "club for growth", "freedom works", "tea party", "heritage action",
    "citizens united", "crossroads", "american crossroads",
    "congressional leadership fund", "senate leadership fund",
]
REBNY_BOILERPLATE_LINES = {
    "member directory",
    "member resources",
    "search by name",
    "search",
    "no search results found",
    "contact us",
    "join us",
    "login",
    "become a member",
    "terms of use",
    "privacy policy",
    "residential listing service rls",
    "webinar hub",
    "nyc lease",
    "access member resources",
    "residential brokerage",
    "commercial brokerage",
    "owners managers",
    "allied associates",
    "events education",
    "upcoming events",
    "rebny awards",
    "sponsorships",
    "about rebny education",
    "online courses",
    "news media",
    "press releases",
    "photos",
    "videos",
    "podcast",
    "style guide",
    "advocacy",
    "research reports",
    "testimony comments",
    "about",
    "organization",
    "committees",
    "rebny foundation",
    "leadership",
    "staff",
    "chair",
    "membership",
    "careers",
    "faq",
    "member disputes",
}


# --------------------------------------------------------------------------- #
# FEC helpers
# --------------------------------------------------------------------------- #
def is_republican_recipient(committee_name, party):
    if party and party.upper() in REPUBLICAN_PARTY_CODES:
        return True
    if committee_name:
        cn = committee_name.lower()
        return any(kw in cn for kw in REPUBLICAN_COMMITTEE_KEYWORDS)
    return False


def lookup_donor(first_name, last_name, api_key):
    name_query = f"{last_name}, {first_name}".upper()
    url = "https://api.open.fec.gov/v1/schedules/schedule_a/"
    params = {
        "contributor_name": name_query,
        "two_year_transaction_period": [2026, 2024, 2022],
        "per_page": 100,
        "api_key": api_key,
        "sort": "-contribution_receipt_date",
    }
    fail = {
        "fec_status": "error", "flag": False, "needs_review": False,
        "total_contributions": 0, "republican_count": 0,
        "republican_total": 0, "top_recipients": "", "fec_detail": "",
    }
    try:
        resp = requests.get(url, params=params, timeout=10)
        if resp.status_code == 429:
            fail.update({"fec_status": "rate_limited", "fec_detail": "Rate limited"})
            return fail
        if resp.status_code != 200:
            fail.update({"fec_status": "error", "fec_detail": f"HTTP {resp.status_code}"})
            return fail

        data = resp.json()
        results = data.get("results", [])
        republican_donations = []

        for r in results:
            committee = r.get("committee", {}) or {}
            committee_name = committee.get("name", "") or r.get("committee_name", "") or ""
            party = committee.get("party", "") or ""
            amount = r.get("contribution_receipt_amount", 0) or 0
            if is_republican_recipient(committee_name, party):
                republican_donations.append({"committee": committee_name, "amount": amount})

        flagged = len(republican_donations) > 0
        rep_total = sum(d["amount"] for d in republican_donations)
        top_recipients = list({d["committee"] for d in republican_donations})[:3]
        total_count = data.get("pagination", {}).get("count", len(results))
        needs_review = total_count >= FALSE_POSITIVE_THRESHOLD

        if needs_review and flagged:
            detail = f"{total_count} total FEC records - common name, verify manually. GOP: ${rep_total:,.0f}"
        elif flagged:
            detail = f"${rep_total:,.0f} across {len(republican_donations)} donation(s)"
        elif needs_review:
            detail = f"{total_count} total FEC records - common name, verify manually"
        else:
            detail = "No Republican donations found"

        return {
            "fec_status": "ok",
            "flag": flagged,
            "needs_review": needs_review,
            "total_contributions": total_count,
            "republican_count": len(republican_donations),
            "republican_total": rep_total,
            "top_recipients": ", ".join(top_recipients) if top_recipients else "",
            "fec_detail": detail,
        }

    except requests.exceptions.Timeout:
        fail.update({"fec_status": "timeout", "fec_detail": "Request timed out"})
        return fail
    except Exception as e:
        fail.update({"fec_detail": str(e)})
        return fail


# --------------------------------------------------------------------------- #
# REBNY parsing helpers
# --------------------------------------------------------------------------- #
def normalize_text(value):
    """Normalize visible text so REBNY matches survive punctuation/case differences."""
    value = unicodedata.normalize("NFKD", str(value))
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    value = value.lower().replace("&", " and ")
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def name_tokens(value):
    return [tok for tok in normalize_text(value).split() if tok]


def target_name_patterns(first_name, last_name):
    """
    Build strict-enough name patterns for directory cards.

    We require the first-name token and every last-name token. Middle initials/names
    may appear between them, and the site may display either "First Last" or
    "Last, First". This intentionally does NOT trust the directory's generic
    "3 Members" count.
    """
    first = name_tokens(first_name)
    last = name_tokens(last_name)
    if not first or not last:
        return []

    first_main = re.escape(first[0])
    last_phrase = r"\s+".join(re.escape(tok) for tok in last)
    spacer = r"(?:\s+[a-z0-9]+){0,4}\s+"
    return [
        re.compile(rf"\b{first_main}\b{spacer}{last_phrase}\b"),
        re.compile(rf"\b{last_phrase}\b{spacer}{first_main}\b"),
    ]


def likely_result_lines(page_text):
    """
    Return visible lines that could plausibly be REBNY result cards.

    The directory page includes navigation, footer text, and sometimes stale count
    text. Filtering boilerplate lines keeps the parser focused on actual results.
    """
    lines = []
    seen = set()
    for raw_line in str(page_text).splitlines():
        display_line = " ".join(raw_line.split()).strip()
        normalized = normalize_text(display_line)
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)

        if normalized in REBNY_BOILERPLATE_LINES:
            continue
        if re.fullmatch(r"\d+\s+members?", normalized):
            continue
        if normalized.startswith("image ") or normalized.startswith("btn arrow"):
            continue
        if "copyright" in normalized or normalized.startswith("202"):
            continue
        if len(normalized) < 3:
            continue

        lines.append(display_line)
    return lines


def find_rebny_name_match(page_text, first_name, last_name):
    """Return the matching visible result line, or None if no exact name result appears."""
    patterns = target_name_patterns(first_name, last_name)
    if not patterns:
        return None

    candidates = likely_result_lines(page_text)

    # First try line-level matches: safest because real cards usually put the
    # person name on one visible line.
    for line in candidates:
        normalized_line = normalize_text(line)
        if any(pattern.search(normalized_line) for pattern in patterns):
            return line

    # Do not merge separate result-card lines. Combining lines can accidentally
    # stitch together two different people and create a false positive.
    return None


def extract_rebny_count(page_text):
    count_match = re.search(r"\b([\d,]+)\s+Members?\b", str(page_text), re.IGNORECASE)
    return int(count_match.group(1).replace(",", "")) if count_match else None


def classify_rebny_page_text(page_text, first_name, last_name):
    """
    Convert searched REBNY page text into the app's result dictionary.

    Rules:
    1. FOUND only when the searched first+last name appears in visible result text.
    2. NOT FOUND when the page explicitly says no results or no count exists.
    3. REVIEW when REBNY shows result count(s), but none are exact name matches.
    """
    query = f"{first_name} {last_name}".strip()
    count = extract_rebny_count(page_text)
    matched_line = find_rebny_name_match(page_text, first_name, last_name)
    no_results = re.search(r"No\s+Search\s+Results", str(page_text), re.IGNORECASE) is not None

    if matched_line:
        detail = "Exact name appears in REBNY directory results"
        if count is not None:
            detail += f" ({count} result(s) shown)"
        detail += f": {matched_line}"
        return {
            "rebny_status": "FOUND",
            "rebny_match": True,
            "rebny_result_count": count if count is not None else "",
            "rebny_detail": detail,
        }

    if no_results or count == 0 or count is None:
        return {
            "rebny_status": "not found",
            "rebny_match": False,
            "rebny_result_count": count if count is not None else 0,
            "rebny_detail": "Not in REBNY directory",
        }

    return {
        "rebny_status": "review",
        "rebny_match": False,
        "rebny_result_count": count,
        "rebny_detail": f"{count} result(s) shown, but no exact visible match for {query}",
    }


@dataclass
class REBNYDirectoryClient:
    """
    Reusable Playwright client for REBNY lookups.

    Launching Chromium once per spreadsheet is faster and more reliable than
    opening a new browser for every donor. Each lookup reloads the directory,
    fills the actual Search By Name field, submits it, then parses visible text.
    """
    headless: bool = True
    browser_timeout_ms: int = 30000
    settle_ms: int = 1200
    playwright: Optional[object] = None
    browser: Optional[object] = None
    page: Optional[object] = None

    def start(self):
        if not PLAYWRIGHT_AVAILABLE:
            raise RuntimeError("Playwright is not installed")
        self.playwright = sync_playwright().start()
        self.browser = self.playwright.chromium.launch(headless=self.headless)
        self.page = self.browser.new_page(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124 Safari/537.36"
            )
        )
        return self

    def close(self):
        try:
            if self.browser:
                self.browser.close()
        finally:
            self.browser = None
            self.page = None
            if self.playwright:
                self.playwright.stop()
                self.playwright = None

    def __enter__(self):
        return self.start()

    def __exit__(self, exc_type, exc, tb):
        self.close()

    def lookup(self, first_name, last_name):
        if not self.page:
            raise RuntimeError("REBNYDirectoryClient.start() must be called first")

        query = f"{first_name} {last_name}".strip()
        default = {
            "rebny_status": "unknown",
            "rebny_match": False,
            "rebny_result_count": "",
            "rebny_detail": "Could not reach REBNY",
        }

        if not query:
            default.update({"rebny_status": "not found", "rebny_detail": "Blank name"})
            return default

        try:
            self.page.goto(REBNY_DIRECTORY_URL, wait_until="domcontentloaded", timeout=self.browser_timeout_ms)
            self._quiet_wait_for_network()
            search_box = self._find_search_input()

            if not search_box:
                default.update({
                    "rebny_status": "parse error",
                    "rebny_detail": "Could not find the REBNY Search By Name input",
                })
                return default

            self._run_search(search_box, query)
            page_text = self._results_text()
            return classify_rebny_page_text(page_text, first_name, last_name)

        except Exception as e:
            default["rebny_detail"] = f"Error: {e}"
            return default

    def _quiet_wait_for_network(self):
        try:
            self.page.wait_for_load_state("networkidle", timeout=10000)
        except Exception:
            pass

    def _first_visible(self, selector):
        locator = self.page.locator(selector)
        try:
            count = min(locator.count(), 30)
        except Exception:
            return None

        for i in range(count):
            item = locator.nth(i)
            try:
                if item.is_visible(timeout=500):
                    return item
            except Exception:
                continue
        return None

    def _find_search_input(self):
        selectors = [
            "main input[type='search']",
            "main input[placeholder*='Search' i]",
            "main input[aria-label*='Search' i]",
            "main input[name*='search' i]",
            "main input[type='text']",
            "input[type='search']",
            "input[placeholder*='Search' i]",
            "input[aria-label*='Search' i]",
            "input[name*='search' i]",
            "input[type='text']",
        ]
        for selector in selectors:
            found = self._first_visible(selector)
            if found:
                return found
        return None

    def _run_search(self, search_box, query):
        search_box.scroll_into_view_if_needed(timeout=5000)
        search_box.click(timeout=5000)
        search_box.fill("", timeout=5000)
        search_box.fill(query, timeout=5000)

        # Some JS widgets listen to both input and change events. Playwright's
        # fill() normally emits them, but this explicit dispatch avoids silent
        # failures on custom search components.
        try:
            handle = search_box.element_handle(timeout=3000)
            self.page.evaluate(
                """
                (el) => {
                    el.dispatchEvent(new Event('input', { bubbles: true }));
                    el.dispatchEvent(new Event('change', { bubbles: true }));
                }
                """,
                handle,
            )
        except Exception:
            pass

        submitted = False
        for selector in [
            "main form button[type='submit']",
            "main form input[type='submit']",
            "main button[type='submit']",
            "main button:has-text('Search')",
            "main [role='button']:has-text('Search')",
            "button[type='submit']",
            "button:has-text('Search')",
            "[role='button']:has-text('Search')",
        ]:
            button = self._first_visible(selector)
            if button:
                try:
                    button.click(timeout=5000)
                    submitted = True
                    break
                except Exception:
                    continue

        if not submitted:
            try:
                search_box.press("Enter", timeout=5000)
                submitted = True
            except Exception:
                pass

        if not submitted:
            # Last-resort form submit. This is intentionally after the normal
            # click/Enter path so React-style handlers still get first chance.
            try:
                handle = search_box.element_handle(timeout=3000)
                self.page.evaluate(
                    """
                    (el) => {
                        if (el.form && el.form.requestSubmit) el.form.requestSubmit();
                        else if (el.form) el.form.submit();
                    }
                    """,
                    handle,
                )
            except Exception:
                pass

        self._quiet_wait_for_network()
        try:
            self.page.wait_for_timeout(self.settle_ms)
        except Exception:
            pass

    def _results_text(self):
        for selector in ["main", "body"]:
            try:
                text = self.page.locator(selector).inner_text(timeout=5000)
                if text and text.strip():
                    return text
            except Exception:
                continue
        try:
            return self.page.content()
        except Exception:
            return ""


def lookup_rebny(first_name, last_name, client=None):
    """Single-name REBNY lookup. The main app passes a reusable client."""
    default = {
        "rebny_status": "unknown",
        "rebny_match": False,
        "rebny_result_count": "",
        "rebny_detail": "Could not reach REBNY",
    }

    if not PLAYWRIGHT_AVAILABLE:
        default["rebny_status"] = "unavailable"
        default["rebny_detail"] = (
            "Playwright not installed. Run: pip install playwright && playwright install chromium"
        )
        return default

    if client:
        return client.lookup(first_name, last_name)

    try:
        with REBNYDirectoryClient() as one_off_client:
            return one_off_client.lookup(first_name, last_name)
    except Exception as e:
        default["rebny_detail"] = f"Error: {e}"
        return default


# --------------------------------------------------------------------------- #
# Status helpers and tests
# --------------------------------------------------------------------------- #
def fec_status_label(r):
    if r.get("needs_review") and r.get("flag"):
        return "FLAGGED - REVIEW"
    if r.get("flag"):
        return "FLAGGED"
    if r.get("needs_review"):
        return "REVIEW NEEDED"
    return "Clean"


def run_self_tests():
    """Fast offline tests for the bug-prone REBNY parsing logic."""
    tests = [
        (
            "generic count plus no-results must not be FOUND",
            "Member Directory\nSearch By Name\n3 Members\nNo Search Results Found",
            "Jane",
            "Definitelyfake",
            "not found",
            False,
        ),
        (
            "exact first-last result is FOUND",
            "Member Directory\n1 Member\nJane Definitelyfake\nAcme Realty\nLicensed Real Estate Salesperson",
            "Jane",
            "Definitelyfake",
            "FOUND",
            True,
        ),
        (
            "last-first result is FOUND",
            "1 Member\nDefinitelyfake, Jane\nAcme Realty",
            "Jane",
            "Definitelyfake",
            "FOUND",
            True,
        ),
        (
            "middle initial still matches",
            "1 Member\nJane Q Definitelyfake\nAcme Realty",
            "Jane",
            "Definitelyfake",
            "FOUND",
            True,
        ),
        (
            "similar last name is not exact",
            "2 Members\nJane Definitelyfakes\nJohn Definitelyfake",
            "Jane",
            "Definitelyfake",
            "review",
            False,
        ),
        (
            "multi-word last name matches",
            "1 Member\nMaria De La Cruz\nBrokerage",
            "Maria",
            "De La Cruz",
            "FOUND",
            True,
        ),
        (
            "accent and punctuation normalization matches",
            "1 Member\nJose O Connor\nBrokerage",
            "José",
            "O'Connor",
            "FOUND",
            True,
        ),
    ]

    for label, page_text, first, last, expected_status, expected_match in tests:
        result = classify_rebny_page_text(page_text, first, last)
        assert result["rebny_status"] == expected_status, (label, result)
        assert result["rebny_match"] is expected_match, (label, result)

    print(f"All {len(tests)} REBNY parser self-tests passed.")


# --------------------------------------------------------------------------- #
# Streamlit app
# --------------------------------------------------------------------------- #
def configure_page():
    if st is None:
        raise RuntimeError("Streamlit is not installed. Run: pip install -r requirements(4).txt")

    st.set_page_config(
        page_title="Donor Vetting Tool",
        page_icon=":mag:",
        layout="centered",
    )

    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600&family=DM+Mono:wght@400;500&display=swap');
    html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }
    .stApp { background-color: #0d1117; color: #e6edf3; }
    h1, h2, h3 { font-family: 'DM Sans', sans-serif; font-weight: 600; }
    .hero { text-align: center; padding: 2.5rem 0 1.5rem; }
    .hero h1 { font-size: 2rem; color: #e6edf3; letter-spacing: -0.5px; margin-bottom: 0.4rem; }
    .hero p { color: #8b949e; font-size: 0.95rem; margin: 0; }
    .badge {
        display: inline-block; background: #1c2a3a; color: #58a6ff;
        border: 1px solid #30363d; border-radius: 20px; padding: 2px 12px;
        font-size: 0.75rem; font-family: 'DM Mono', monospace; margin-bottom: 1rem;
    }
    .card { background: #161b22; border: 1px solid #30363d; border-radius: 10px; padding: 1.5rem; margin: 1rem 0; }
    .stat-row { display: flex; gap: 1rem; margin: 1rem 0; flex-wrap: wrap; }
    .stat { flex: 1; min-width: 120px; background: #0d1117; border: 1px solid #30363d; border-radius: 8px; padding: 1rem; text-align: center; }
    .stat .num { font-size: 1.8rem; font-weight: 600; font-family: 'DM Mono', monospace; }
    .stat .label { font-size: 0.75rem; color: #8b949e; margin-top: 2px; text-transform: uppercase; letter-spacing: 0.5px; }
    .flagged { color: #f85149; }
    .clean { color: #3fb950; }
    .info-box {
        background: #1c2a3a; border-left: 3px solid #388bfd;
        border-radius: 0 6px 6px 0; padding: 0.75rem 1rem;
        font-size: 0.85rem; color: #8b949e; margin: 0.75rem 0;
    }
    </style>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div class="hero">
        <div class="badge">FEC &middot; OpenFEC API &middot; REBNY</div>
        <h1>&#128269; Donor Vetting Tool</h1>
        <p>Upload a spreadsheet of names &mdash; check FEC records for Republican donations<br>and look up REBNY membership.</p>
    </div>
    """, unsafe_allow_html=True)


def read_uploaded_names(uploaded):
    df = pd.read_csv(uploaded) if uploaded.name.endswith(".csv") else pd.read_excel(uploaded)
    df.columns = [c.strip().lower() for c in df.columns]

    col_map = {}
    for c in df.columns:
        if "first" in c and "first" not in col_map:
            col_map["first"] = c
        if "last" in c and "last" not in col_map:
            col_map["last"] = c

    if "first" not in col_map or "last" not in col_map:
        raise ValueError("Could not find 'First Name' and 'Last Name' columns in your file.")

    df_names = df[[col_map["first"], col_map["last"]]].copy()
    df_names.columns = ["First Name", "Last Name"]
    df_names = df_names.fillna("")
    return df_names


def build_excel_export(uploaded, df_names, results_list, run_fec, include_rebny):
    uploaded.seek(0)
    if uploaded.name.endswith(".xlsx"):
        wb = load_workbook(uploaded)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(list(df_names.columns))
        for row in df_names.itertuples(index=False):
            ws.append(list(row))

    flag_col = ws.max_column + 1
    export_headers = []
    if run_fec:
        export_headers += ["FEC Status", "GOP Donations ($)", "Top Recipients", "FEC Records", "FEC Detail"]
    if include_rebny:
        export_headers += ["REBNY Member?", "REBNY Result Count", "REBNY Detail"]

    for i, h in enumerate(export_headers):
        cell = ws.cell(1, flag_col + i)
        cell.value = h
        cell.font = Font(bold=True)

    red_fill = PatternFill("solid", start_color="FFCCCC")
    orange_fill = PatternFill("solid", start_color="FFE5B4")
    green_fill = PatternFill("solid", start_color="CCFFCC")
    purple_fill = PatternFill("solid", start_color="E8D5FF")

    for i, r in enumerate(results_list, start=2):
        offset = 0
        if run_fec:
            ws.cell(i, flag_col + offset).value = fec_status_label(r)
            ws.cell(i, flag_col + offset + 1).value = f"${r['republican_total']:,.0f}" if r.get("flag") else ""
            ws.cell(i, flag_col + offset + 2).value = r.get("top_recipients", "")
            ws.cell(i, flag_col + offset + 3).value = r.get("total_contributions", 0)
            ws.cell(i, flag_col + offset + 4).value = r.get("fec_detail", "")
            offset += 5

        if include_rebny:
            ws.cell(i, flag_col + offset).value = "FOUND" if r.get("rebny_match") else r.get("rebny_status", "")
            ws.cell(i, flag_col + offset + 1).value = r.get("rebny_result_count", "")
            ws.cell(i, flag_col + offset + 2).value = r.get("rebny_detail", "")

        if run_fec and r.get("flag") and not r.get("needs_review"):
            fill = red_fill
        elif run_fec and r.get("needs_review"):
            fill = orange_fill
        elif include_rebny and r.get("rebny_match"):
            fill = purple_fill
        else:
            fill = green_fill

        for c in range(1, flag_col + len(export_headers)):
            ws.cell(i, c).fill = fill

    for c in range(flag_col, flag_col + len(export_headers)):
        ws.column_dimensions[get_column_letter(c)].width = 26

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def main():
    configure_page()

    with st.expander("API Settings", expanded=False):
        api_key_input = st.text_input(
            "OpenFEC API Key",
            type="password",
            placeholder="Paste your key from api.open.fec.gov/developers",
            help="Free key - 1,000 requests/hour.",
        )
        st.markdown(
            '<div class="info-box">Your key is sent directly to the FEC API and never stored.</div>',
            unsafe_allow_html=True,
        )

    api_key = api_key_input.strip() if api_key_input else "DEMO_KEY"

    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown("### Upload Your Spreadsheet")
    st.markdown(
        '<div class="info-box">File must have <b>First Name</b> and <b>Last Name</b> columns. CSV or XLSX accepted.</div>',
        unsafe_allow_html=True,
    )
    uploaded = st.file_uploader("", type=["xlsx", "csv"], label_visibility="collapsed")
    st.markdown("</div>", unsafe_allow_html=True)

    if not uploaded:
        return

    try:
        df_names = read_uploaded_names(uploaded)
    except Exception as e:
        st.error(str(e))
        st.stop()

    st.markdown(f"**{len(df_names)} names loaded.**")
    st.dataframe(df_names.head(5), use_container_width=True, hide_index=True)

    st.markdown("**Checks to run:**")
    run_fec = st.checkbox("FEC Republican donation check", value=True)
    run_rebny = st.checkbox("REBNY member directory check", value=True)

    with st.expander("How the REBNY match works", expanded=False):
        st.markdown(
            "The app opens the REBNY Member Directory, types each person into the visible **Search By Name** field, "
            "submits the search, and only marks **FOUND** when that exact first + last name appears in the visible results. "
            "Generic text such as `3 Members` is never enough by itself."
        )

    if not PLAYWRIGHT_AVAILABLE and run_rebny:
        st.warning(
            "Playwright is not installed - REBNY check will be skipped. "
            "Run `pip install playwright && playwright install chromium` then restart Streamlit."
        )

    if not st.button("Run Vetting", type="primary", use_container_width=True):
        return

    if not run_fec and not run_rebny:
        st.warning("Select at least one check to run.")
        st.stop()

    include_rebny = run_rebny and PLAYWRIGHT_AVAILABLE
    results_list = []
    progress = st.progress(0, text="Starting...")
    status_box = st.empty()
    total = len(df_names)
    rebny_client = None

    try:
        if include_rebny:
            status_box.markdown(
                '<div class="info-box">Starting REBNY browser session...</div>',
                unsafe_allow_html=True,
            )
            try:
                rebny_client = REBNYDirectoryClient().start()
            except Exception as e:
                st.error(
                    "Could not start the REBNY browser session. "
                    "Run `playwright install chromium` after installing requirements, then restart Streamlit. "
                    f"Details: {e}"
                )
                st.stop()

        for position, row in enumerate(df_names.itertuples(index=False), start=1):
            first = str(getattr(row, "_0", row[0])).strip()
            last = str(getattr(row, "_1", row[1])).strip()
            status_box.markdown(
                f'<div class="info-box">Checking <b>{first} {last}</b> ({position} of {total})...</div>',
                unsafe_allow_html=True,
            )

            result = {"First Name": first, "Last Name": last}

            if run_fec:
                result.update(lookup_donor(first, last, api_key))
                time.sleep(0.15)

            if include_rebny:
                result.update(lookup_rebny(first, last, client=rebny_client))
                time.sleep(0.1)

            results_list.append(result)
            progress.progress(position / total, text=f"{position}/{total} checked")

    finally:
        if rebny_client:
            rebny_client.close()
        status_box.empty()
        progress.empty()

    out_df = df_names.copy()

    if run_fec:
        out_df["FEC Status"] = [fec_status_label(r) for r in results_list]
        out_df["GOP Donations ($)"] = [
            f"${r['republican_total']:,.0f}" if r.get("flag") else "-"
            for r in results_list
        ]
        out_df["Top Recipients"] = [r.get("top_recipients", "") for r in results_list]
        out_df["FEC Records"] = [r.get("total_contributions", 0) for r in results_list]
        out_df["FEC Detail"] = [r.get("fec_detail", "") for r in results_list]

    if include_rebny:
        out_df["REBNY Member?"] = [
            "YES" if r.get("rebny_match") else r.get("rebny_status", "unknown")
            for r in results_list
        ]
        out_df["REBNY Result Count"] = [r.get("rebny_result_count", "") for r in results_list]
        out_df["REBNY Detail"] = [r.get("rebny_detail", "") for r in results_list]

    stat_html = (
        f'<div class="stat-row">'
        f'<div class="stat"><div class="num">{total}</div>'
        f'<div class="label">Total Checked</div></div>'
    )

    if run_fec:
        flagged_count = sum(1 for r in results_list if r.get("flag") and not r.get("needs_review"))
        review_count = sum(1 for r in results_list if r.get("needs_review"))
        clean_count = total - flagged_count - review_count
        stat_html += (
            f'<div class="stat"><div class="num flagged">{flagged_count}</div>'
            f'<div class="label">FEC Flagged</div></div>'
            f'<div class="stat"><div class="num" style="color:#e3b341">{review_count}</div>'
            f'<div class="label">FEC Review</div></div>'
            f'<div class="stat"><div class="num clean">{clean_count}</div>'
            f'<div class="label">FEC Clean</div></div>'
        )

    if include_rebny:
        rebny_found = sum(1 for r in results_list if r.get("rebny_match"))
        rebny_review = sum(1 for r in results_list if r.get("rebny_status") == "review")
        stat_html += (
            f'<div class="stat"><div class="num" style="color:#d2a8ff">{rebny_found}</div>'
            f'<div class="label">REBNY Found</div></div>'
            f'<div class="stat"><div class="num" style="color:#e3b341">{rebny_review}</div>'
            f'<div class="label">REBNY Review</div></div>'
        )

    stat_html += "</div>"
    st.markdown(stat_html, unsafe_allow_html=True)

    st.markdown("### Results")
    st.dataframe(out_df, use_container_width=True, hide_index=True)

    buf = build_excel_export(uploaded, df_names, results_list, run_fec, include_rebny)
    base_name = uploaded.name.rsplit(".", 1)[0]
    st.download_button(
        label="Download Vetted Spreadsheet",
        data=buf,
        file_name=f"{base_name}_vetted.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        type="primary",
    )


if __name__ == "__main__":
    if "--self-test" in sys.argv:
        run_self_tests()
    else:
        main()
