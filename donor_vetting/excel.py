"""Excel export helpers for vetted results."""
from __future__ import annotations

import io
from typing import Any

FILL_COLORS = {
    "flag": "FFCCCC",
    "review": "FFE5B4",
    "rebny": "E8D5FF",
    "clean": "CCFFCC",
}


def dataframe_to_excel_bytes(df: Any) -> bytes:
    """Return an xlsx workbook as bytes with simple status coloring."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Vetted Results"

    headers = list(df.columns)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    fills = {key: PatternFill("solid", start_color=color) for key, color in FILL_COLORS.items()}
    for _, row in df.iterrows():
        values = [row.get(header, "") for header in headers]
        ws.append(values)
        excel_row = ws.max_row
        fec_status = str(row.get("FEC Status", "")).lower()
        rebny_status = str(row.get("REBNY Status", "")).lower()
        if "flagged" in fec_status:
            fill = fills["flag"]
        elif "review" in fec_status or "review" in rebny_status:
            fill = fills["review"]
        elif rebny_status == "found":
            fill = fills["rebny"]
        else:
            fill = fills["clean"]
        for col in range(1, len(headers) + 1):
            ws.cell(excel_row, col).fill = fill

    for index, header in enumerate(headers, start=1):
        max_len = max([len(str(header))] + [len(str(ws.cell(row=r, column=index).value or "")) for r in range(2, ws.max_row + 1)])
        ws.column_dimensions[get_column_letter(index)].width = min(max(12, max_len + 2), 48)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
