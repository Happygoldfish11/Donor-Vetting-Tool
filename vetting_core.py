from __future__ import annotations

import io
import math
import re
import time
import unicodedata
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Iterable, Optional

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

DEFAULT_REBNY_CACHE_PATH = Path("data/rebny_members.xlsx")
DEFAULT_FEC_YEARS = [2026, 2024, 2022]
FEC_API_URL = "https://api.open.fec.gov/v1/schedules/schedule_a/"

FIRST_NAME_ALIASES = ["first name", "firstname", "first", "given name", "given"]
LAST_NAME_ALIASES = ["last name", "lastname", "last", "surname", "family name", "family"]
STATE_ALIASES = ["state", "st"]
ZIP_ALIASES = ["zip", "zipcode", "zip code", "postal", "postal code"]

REBNY_NAME_ALIASES = [
    "name",
    "member name",
    "full name",
    "fullname",
    "person",
    "member",
    "contact",
]
REBNY_COMPANY_ALIASES = ["company", "firm", "organization", "org", "brokerage", "office"]
REBNY_TITLE_ALIASES = ["title", "role", "position", "job title"]
REBNY_URL_ALIASES = ["url", "profile", "profile url", "link", "source url"]

NAME_SUFFIXES = {
    "jr",
    "sr",
    "ii",
    "iii",
    "iv",
    "v",
    "vi",
    "esq",
    "phd",
    "md",
    "mba",
    "cpa",
}

REPUBLICAN_PARTY_CODES = {"REP", "R"}
REPUBLICAN_COMMITTEE_TERMS = [
    "republican",
    "gop",
    "rnc",
    "nrcc",
    "nrsc",
    "maga",
    "make america great again",
    "trump",
    "desantis",
    "haley for president",
    "ted cruz",
    "cruz for",
    "cotton for",
    "tim scott",
    "mike pence",
    "right to rise",
    "club for growth",
    "freedomworks",
    "freedom works",
    "heritage action",
    "tea party",
    "american crossroads",
    "crossroads gps",
    "congressional leadership fund",
    "senate leadership fund",
    "conservative victory",
    "conservative pac",
]


@dataclass(slots=True)
class Person:
    first_name: str
    last_name: str
    state: str = ""
    zip_code: str = ""
    original_index: int = 0

    @property
    def full_name(self) -> str:
        return f"{self.first_name} {self.last_name}".strip()


@dataclass(slots=True)
class RebnyMember:
    name: str
    first_name: str = ""
    last_name: str = ""
    company: str = ""
    title: str = ""
    profile_url: str = ""
    raw_text: str = ""

    @property
    def display_name(self) -> str:
        return self.name or f"{self.first_name} {self.last_name}".strip()


@dataclass(slots=True)
class RebnyResult:
    status: str
    result: str
    found: bool
    review: bool
    match_name: str = ""
    company: str = ""
    score: int = 0
    detail: str = ""

    def as_row(self) -> dict[str, Any]:
        return {
            "REBNY Status": self.status,
            "REBNY Result": self.result,
            "REBNY Match Name": self.match_name,
            "REBNY Company": self.company,
            "REBNY Score": self.score,
            "REBNY Detail": self.detail,
        }


@dataclass(slots=True)
class FecResult:
    status: str
    result: str
    flagged: bool
    review: bool
    republican_total: float = 0.0
    republican_count: int = 0
    total_records: int = 0
    top_recipients: str = ""
    detail: str = ""

    def as_row(self) -> dict[str, Any]:
        return {
            "FEC Status": self.status,
            "FEC Result": self.result,
            "GOP Donations ($)": round(self.republican_total, 2),
            "GOP Donation Count": self.republican_count,
            "FEC Records Checked": self.total_records,
            "Top GOP Recipients": self.top_recipients,
            "FEC Detail": self.detail,
        }


def clean_scalar(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    text = str(value).strip()
    if text.lower() in {"nan", "none", "null"}:
        return ""
    return text


def ascii_fold(text: str) -> str:
    text = unicodedata.normalize("NFKD", clean_scalar(text))
    return "".join(ch for ch in text if not unicodedata.combining(ch))


def normalize_text(text: str) -> str:
    text = ascii_fold(text).lower()
    text = text.replace("&", " and ")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def name_tokens(text: str) -> list[str]:
    tokens = normalize_text(text).split()
    return [token for token in tokens if token and token not in NAME_SUFFIXES]


def normalize_zip(value: Any) -> str:
    text = clean_scalar(value)
    digits = re.sub(r"\D", "", text)
    return digits[:5]


def canonical_col(col: Any) -> str:
    return normalize_text(str(col))


def find_column(columns: Iterable[Any], aliases: Iterable[str]) -> Optional[Any]:
    alias_set = {canonical_col(alias) for alias in aliases}
    normalized = {col: canonical_col(col) for col in columns}

    for col, canon in normalized.items():
        if canon in alias_set:
            return col

    for col, canon in normalized.items():
        if any(alias in canon for alias in alias_set):
            return col

    return None


def read_spreadsheet(file_or_path: Any) -> pd.DataFrame:
    name = getattr(file_or_path, "name", str(file_or_path)).lower()
    if name.endswith(".csv"):
        return pd.read_csv(file_or_path)
    return pd.read_excel(file_or_path)


def people_from_dataframe(df: pd.DataFrame) -> tuple[list[Person], dict[str, Any]]:
    first_col = find_column(df.columns, FIRST_NAME_ALIASES)
    last_col = find_column(df.columns, LAST_NAME_ALIASES)
    state_col = find_column(df.columns, STATE_ALIASES)
    zip_col = find_column(df.columns, ZIP_ALIASES)

    if first_col is None or last_col is None:
        raise ValueError("Your file needs First Name and Last Name columns.")

    people: list[Person] = []
    for idx, row in df.iterrows():
        first = clean_scalar(row.get(first_col, ""))
        last = clean_scalar(row.get(last_col, ""))
        if not first and not last:
            continue
        people.append(
            Person(
                first_name=first,
                last_name=last,
                state=clean_scalar(row.get(state_col, "")) if state_col is not None else "",
                zip_code=normalize_zip(row.get(zip_col, "")) if zip_col is not None else "",
                original_index=int(idx),
            )
        )

    return people, {
        "first": first_col,
        "last": last_col,
        "state": state_col,
        "zip": zip_col,
    }


def people_preview_rows(people: list[Person], limit: int = 10) -> pd.DataFrame:
    rows = [
        {
            "First Name": person.first_name,
            "Last Name": person.last_name,
            "State": person.state,
            "Zip": person.zip_code,
        }
        for person in people[:limit]
    ]
    return pd.DataFrame(rows)


def member_from_row(row: pd.Series, columns: Iterable[Any]) -> RebnyMember:
    name_col = find_column(columns, REBNY_NAME_ALIASES)
    first_col = find_column(columns, FIRST_NAME_ALIASES)
    last_col = find_column(columns, LAST_NAME_ALIASES)
    company_col = find_column(columns, REBNY_COMPANY_ALIASES)
    title_col = find_column(columns, REBNY_TITLE_ALIASES)
    url_col = find_column(columns, REBNY_URL_ALIASES)

    first = clean_scalar(row.get(first_col, "")) if first_col is not None else ""
    last = clean_scalar(row.get(last_col, "")) if last_col is not None else ""
    name = clean_scalar(row.get(name_col, "")) if name_col is not None else ""
    if not name:
        name = f"{first} {last}".strip()

    return RebnyMember(
        name=name,
        first_name=first,
        last_name=last,
        company=clean_scalar(row.get(company_col, "")) if company_col is not None else "",
        title=clean_scalar(row.get(title_col, "")) if title_col is not None else "",
        profile_url=clean_scalar(row.get(url_col, "")) if url_col is not None else "",
        raw_text=" | ".join(clean_scalar(row.get(col, "")) for col in columns if clean_scalar(row.get(col, ""))),
    )


def load_rebny_members(file_or_path: Any = DEFAULT_REBNY_CACHE_PATH) -> list[RebnyMember]:
    if isinstance(file_or_path, (str, Path)) and not Path(file_or_path).exists():
        return []

    df = read_spreadsheet(file_or_path)
    if df.empty:
        return []

    members: list[RebnyMember] = []
    for _, row in df.iterrows():
        member = member_from_row(row, df.columns)
        if member.display_name:
            members.append(member)
    return dedupe_members(members)


def dedupe_members(members: Iterable[RebnyMember]) -> list[RebnyMember]:
    seen: set[tuple[str, str]] = set()
    unique: list[RebnyMember] = []
    for member in members:
        key = (normalize_text(member.display_name), normalize_text(member.company))
        if not key[0] or key in seen:
            continue
        seen.add(key)
        unique.append(member)
    return unique


def parse_member_name(member: RebnyMember) -> tuple[str, str, list[str]]:
    first = normalize_text(member.first_name)
    last = normalize_text(member.last_name)
    tokens = name_tokens(member.display_name)

    if not first and tokens:
        first = tokens[0]
    if not last and len(tokens) >= 2:
        last = tokens[-1]
    return first, last, tokens


def score_rebny_member(person: Person, member: RebnyMember) -> tuple[int, str]:
    query_first = normalize_text(person.first_name)
    query_last = normalize_text(person.last_name)
    member_first, member_last, tokens = parse_member_name(member)
    member_full = normalize_text(member.display_name)
    query_full = normalize_text(person.full_name)

    if not query_first or not query_last or not member_full:
        return 0, "missing name data"

    if member_first == query_first and member_last == query_last:
        return 100, "first and last name exact match"

    if query_first in tokens and query_last in tokens:
        return 98, "first and last tokens found in member name"

    if member_last == query_last and member_first == query_first:
        return 96, "normalized first and last match"

    if member_last == query_last and member_first and query_first and member_first[0] == query_first[0]:
        return 82, "same last name and first initial"

    if member_last == query_last and SequenceMatcher(None, member_first, query_first).ratio() >= 0.87:
        return 80, "same last name and similar first name"

    if SequenceMatcher(None, member_full, query_full).ratio() >= 0.94:
        return 78, "very similar full name"

    if member_last == query_last:
        return 55, "same last name only"

    return 0, "no meaningful name match"


def lookup_rebny_from_members(person: Person, members: list[RebnyMember]) -> RebnyResult:
    if not members:
        return RebnyResult(
            status="cache missing",
            result="REVIEW",
            found=False,
            review=True,
            detail="No REBNY cache loaded. Add data/rebny_members.xlsx or upload a REBNY cache file.",
        )

    scored: list[tuple[int, str, RebnyMember]] = []
    for member in members:
        score, reason = score_rebny_member(person, member)
        if score >= 55:
            scored.append((score, reason, member))

    if not scored:
        return RebnyResult(
            status="not found",
            result="Clean",
            found=False,
            review=False,
            detail="No local REBNY cache match.",
        )

    scored.sort(key=lambda item: item[0], reverse=True)
    best_score, reason, best = scored[0]
    display = best.display_name

    if best_score >= 96:
        return RebnyResult(
            status="found",
            result="FOUND",
            found=True,
            review=False,
            match_name=display,
            company=best.company,
            score=best_score,
            detail=reason,
        )

    if best_score >= 78:
        return RebnyResult(
            status="possible match",
            result="REVIEW",
            found=False,
            review=True,
            match_name=display,
            company=best.company,
            score=best_score,
            detail=f"Possible REBNY match: {reason}",
        )

    nearby = ", ".join(member.display_name for _, _, member in scored[:3])
    return RebnyResult(
        status="same last name",
        result="REVIEW",
        found=False,
        review=True,
        match_name=display,
        company=best.company,
        score=best_score,
        detail=f"Same-last-name cache hit(s), verify manually: {nearby}",
    )


def is_republican_recipient(committee_name: str, party: str = "") -> bool:
    party_norm = normalize_text(party).upper()
    if party_norm in REPUBLICAN_PARTY_CODES:
        return True

    committee_norm = normalize_text(committee_name)
    if not committee_norm:
        return False
    return any(term in committee_norm for term in REPUBLICAN_COMMITTEE_TERMS)


def fec_record_matches_person(record: dict[str, Any], person: Person) -> bool:
    record_name = normalize_text(record.get("contributor_name", ""))
    first = normalize_text(person.first_name)
    last = normalize_text(person.last_name)
    if first and first not in record_name:
        return False
    if last and last not in record_name:
        return False

    if person.state:
        record_state = normalize_text(record.get("contributor_state", ""))
        if record_state and normalize_text(person.state) != record_state:
            return False

    if person.zip_code:
        record_zip = normalize_zip(record.get("contributor_zip", ""))
        if record_zip and person.zip_code != record_zip:
            return False

    return True


def committee_name_and_party(record: dict[str, Any]) -> tuple[str, str]:
    committee = record.get("committee") or {}
    if not isinstance(committee, dict):
        committee = {}
    committee_name = (
        clean_scalar(record.get("committee_name"))
        or clean_scalar(committee.get("name"))
        or clean_scalar(record.get("committee", ""))
    )
    party = (
        clean_scalar(record.get("committee_party"))
        or clean_scalar(record.get("recipient_committee_party"))
        or clean_scalar(committee.get("party"))
        or clean_scalar(committee.get("party_full"))
    )
    return committee_name, party


def lookup_fec(
    person: Person,
    api_key: str,
    years: Iterable[int] = DEFAULT_FEC_YEARS,
    max_pages: int = 5,
    pause_seconds: float = 0.1,
    session: Optional[requests.Session] = None,
) -> FecResult:
    if not api_key:
        api_key = "DEMO_KEY"

    http = session or requests.Session()
    base_params: dict[str, Any] = {
        "api_key": api_key,
        "contributor_name": f"{person.last_name}, {person.first_name}",
        "two_year_transaction_period": list(years),
        "per_page": 100,
        "sort": "-contribution_receipt_date",
    }
    if person.state:
        base_params["contributor_state"] = person.state.upper().strip()
    if person.zip_code:
        base_params["contributor_zip"] = person.zip_code

    records_checked = 0
    republican_donations: list[dict[str, Any]] = []

    try:
        for page in range(1, max_pages + 1):
            params = dict(base_params)
            params["page"] = page
            response = http.get(FEC_API_URL, params=params, timeout=20)

            if response.status_code == 429:
                return FecResult(
                    status="rate limited",
                    result="REVIEW",
                    flagged=False,
                    review=True,
                    total_records=records_checked,
                    detail="FEC rate limit hit. Try again later or use a personal OpenFEC API key.",
                )
            if response.status_code != 200:
                return FecResult(
                    status="error",
                    result="REVIEW",
                    flagged=False,
                    review=True,
                    total_records=records_checked,
                    detail=f"FEC returned HTTP {response.status_code}.",
                )

            payload = response.json()
            results = payload.get("results", [])
            if not results:
                break

            for record in results:
                if not fec_record_matches_person(record, person):
                    continue
                records_checked += 1
                committee_name, party = committee_name_and_party(record)
                if is_republican_recipient(committee_name, party):
                    republican_donations.append(
                        {
                            "committee": committee_name,
                            "amount": float(record.get("contribution_receipt_amount") or 0),
                            "date": clean_scalar(record.get("contribution_receipt_date")),
                            "party": party,
                        }
                    )

            pagination = payload.get("pagination", {}) or {}
            pages = int(pagination.get("pages", page) or page)
            if page >= pages:
                break
            if pause_seconds:
                time.sleep(pause_seconds)

    except requests.exceptions.Timeout:
        return FecResult(
            status="timeout",
            result="REVIEW",
            flagged=False,
            review=True,
            total_records=records_checked,
            detail="FEC request timed out.",
        )
    except Exception as exc:
        return FecResult(
            status="error",
            result="REVIEW",
            flagged=False,
            review=True,
            total_records=records_checked,
            detail=f"FEC lookup failed: {exc}",
        )

    if not republican_donations:
        if records_checked >= 25 and not (person.state or person.zip_code):
            return FecResult(
                status="common name",
                result="REVIEW",
                flagged=False,
                review=True,
                total_records=records_checked,
                detail="Many FEC records matched this name. Add State or Zip to reduce false positives.",
            )
        return FecResult(
            status="clean",
            result="Clean",
            flagged=False,
            review=False,
            total_records=records_checked,
            detail="No Republican recipient matches found in checked FEC records.",
        )

    total = sum(donation["amount"] for donation in republican_donations)
    committees: list[str] = []
    for donation in republican_donations:
        committee = donation["committee"]
        if committee and committee not in committees:
            committees.append(committee)

    result_label = "FLAGGED"
    review = False
    status = "flagged"
    detail = f"${total:,.0f} across {len(republican_donations)} Republican/aligned donation record(s)."
    if records_checked >= 25 and not (person.state or person.zip_code):
        result_label = "FLAGGED — REVIEW"
        review = True
        status = "flagged common name"
        detail += " Common name; verify identity manually."

    return FecResult(
        status=status,
        result=result_label,
        flagged=True,
        review=review,
        republican_total=total,
        republican_count=len(republican_donations),
        total_records=records_checked,
        top_recipients=", ".join(committees[:5]),
        detail=detail,
    )


def build_results_dataframe(
    source_df: pd.DataFrame,
    people: list[Person],
    column_map: dict[str, Any],
    rebny_results: Optional[dict[int, RebnyResult]] = None,
    fec_results: Optional[dict[int, FecResult]] = None,
) -> pd.DataFrame:
    output = source_df.copy()

    for person in people:
        idx = person.original_index
        if rebny_results and idx in rebny_results:
            for col, value in rebny_results[idx].as_row().items():
                output.loc[idx, col] = value
        if fec_results and idx in fec_results:
            for col, value in fec_results[idx].as_row().items():
                output.loc[idx, col] = value

    return output


def results_summary(
    people: list[Person],
    rebny_results: Optional[dict[int, RebnyResult]] = None,
    fec_results: Optional[dict[int, FecResult]] = None,
) -> dict[str, int]:
    summary = {"total": len(people)}
    if rebny_results is not None:
        summary["rebny_found"] = sum(1 for result in rebny_results.values() if result.found)
        summary["rebny_review"] = sum(1 for result in rebny_results.values() if result.review)
    if fec_results is not None:
        summary["fec_flagged"] = sum(1 for result in fec_results.values() if result.flagged)
        summary["fec_review"] = sum(1 for result in fec_results.values() if result.review)
    return summary


def dataframe_to_xlsx_bytes(df: pd.DataFrame) -> bytes:
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Vetting Results")
    buffer.seek(0)

    workbook = load_workbook(buffer)
    sheet = workbook.active
    sheet.freeze_panes = "A2"
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    green = PatternFill("solid", fgColor="D9EAD3")
    red = PatternFill("solid", fgColor="F4CCCC")
    orange = PatternFill("solid", fgColor="FCE5CD")
    purple = PatternFill("solid", fgColor="D9D2E9")

    headers = [cell.value for cell in sheet[1]]
    fec_result_col = headers.index("FEC Result") + 1 if "FEC Result" in headers else None
    rebny_result_col = headers.index("REBNY Result") + 1 if "REBNY Result" in headers else None

    for row in range(2, sheet.max_row + 1):
        fill = green
        fec_value = clean_scalar(sheet.cell(row, fec_result_col).value) if fec_result_col else ""
        rebny_value = clean_scalar(sheet.cell(row, rebny_result_col).value) if rebny_result_col else ""

        if "FLAGGED" in fec_value:
            fill = red
        elif "REVIEW" in fec_value or "REVIEW" in rebny_value:
            fill = orange
        elif "FOUND" in rebny_value:
            fill = purple

        for col in range(1, sheet.max_column + 1):
            sheet.cell(row, col).fill = fill

    for col_idx in range(1, sheet.max_column + 1):
        letter = get_column_letter(col_idx)
        width = min(42, max(12, len(clean_scalar(sheet.cell(1, col_idx).value)) + 4))
        sheet.column_dimensions[letter].width = width

    out = io.BytesIO()
    workbook.save(out)
    return out.getvalue()
